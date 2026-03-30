"""Tests for omopy.testing — test data generation for OMOP CDM studies."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import polars as pl
import pytest

from omopy.generics import CdmReference, CdmSchema, CdmTable, CdmVersion, CohortTable
from omopy.testing import (
    generate_test_tables,
    graph_cohort,
    mock_test_cdm,
    patients_cdm,
    read_patients,
    validate_patient_data,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _minimal_person_df() -> pl.DataFrame:
    """Minimal valid person table."""
    return pl.DataFrame({
        "person_id": [1, 2],
        "gender_concept_id": [8507, 8532],
        "year_of_birth": [1990, 1985],
        "race_concept_id": [8515, 8516],
        "ethnicity_concept_id": [38003563, 38003564],
    })


def _minimal_obs_period_df() -> pl.DataFrame:
    """Minimal valid observation_period table."""
    return pl.DataFrame({
        "observation_period_id": [1, 2],
        "person_id": [1, 2],
        "observation_period_start_date": [date(2020, 1, 1), date(2020, 1, 1)],
        "observation_period_end_date": [date(2024, 12, 31), date(2024, 12, 31)],
        "period_type_concept_id": [44814724, 44814724],
    })


def _minimal_condition_df() -> pl.DataFrame:
    """Minimal valid condition_occurrence table."""
    return pl.DataFrame({
        "condition_occurrence_id": [1],
        "person_id": [1],
        "condition_concept_id": [31967],
        "condition_start_date": [date(2021, 3, 15)],
        "condition_type_concept_id": [32020],
    })


def _cohort_df() -> pl.DataFrame:
    """Cohort table for plotting tests."""
    return pl.DataFrame({
        "cohort_definition_id": [1, 1, 2],
        "subject_id": [1, 2, 1],
        "cohort_start_date": [date(2021, 1, 1), date(2021, 6, 1), date(2022, 1, 1)],
        "cohort_end_date": [date(2021, 12, 31), date(2021, 12, 31), date(2022, 6, 30)],
    })


def _write_test_xlsx(path: Path, tables: dict[str, pl.DataFrame]) -> None:
    """Write an Excel workbook programmatically for tests."""
    import openpyxl

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
    for name, df in tables.items():
        ws = wb.create_sheet(title=name)
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data
        for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
            for col_idx, col_name in enumerate(df.columns, start=1):
                val = row[col_name]
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=val)
    wb.save(str(path))
    wb.close()


def _write_test_json(path: Path, tables: dict[str, pl.DataFrame]) -> None:
    """Write a JSON file for patients_cdm tests."""
    json_data: dict[str, object] = {
        "_meta": {"test_name": "unit_test", "cdm_version": "5.4"},
    }
    for name, df in tables.items():
        rows = []
        for row in df.iter_rows(named=True):
            converted = {}
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    converted[k] = v.isoformat()
                else:
                    converted[k] = v
            rows.append(converted)
        json_data[name] = rows
    path.write_text(json.dumps(json_data, indent=2), encoding="utf-8")


# ===========================================================================
# validate_patient_data
# ===========================================================================


class TestValidatePatientData:
    """Tests for validate_patient_data()."""

    def test_valid_data_returns_empty(self):
        data = {
            "person": _minimal_person_df(),
            "observation_period": _minimal_obs_period_df(),
        }
        errors = validate_patient_data(data, cdm_version="5.4")
        assert errors == []

    def test_unknown_table_returns_error(self):
        data = {"not_a_real_table": pl.DataFrame({"x": [1]})}
        errors = validate_patient_data(data, cdm_version="5.4")
        assert len(errors) == 1
        assert "Unknown CDM table" in errors[0]
        assert "not_a_real_table" in errors[0]

    def test_missing_required_column(self):
        # person without person_id
        data = {
            "person": pl.DataFrame({
                "gender_concept_id": [8507],
                "year_of_birth": [1990],
            }),
        }
        errors = validate_patient_data(data, cdm_version="5.4")
        assert any("person_id" in e for e in errors)

    def test_multiple_missing_columns(self):
        # person with only one column
        data = {"person": pl.DataFrame({"person_id": [1]})}
        errors = validate_patient_data(data, cdm_version="5.4")
        # Should flag gender_concept_id, year_of_birth, etc.
        assert len(errors) >= 3

    def test_extra_columns_accepted(self):
        """Extra columns beyond the spec should not cause errors."""
        df = _minimal_person_df().with_columns(pl.lit("extra_val").alias("my_custom_col"))
        data = {"person": df}
        errors = validate_patient_data(data, cdm_version="5.4")
        assert errors == []

    def test_valid_condition_occurrence(self):
        data = {"condition_occurrence": _minimal_condition_df()}
        errors = validate_patient_data(data, cdm_version="5.4")
        assert errors == []

    def test_version_53(self):
        data = {"person": _minimal_person_df()}
        errors = validate_patient_data(data, cdm_version="5.3")
        assert errors == []

    def test_version_54(self):
        data = {"person": _minimal_person_df()}
        errors = validate_patient_data(data, cdm_version="5.4")
        assert errors == []

    def test_invalid_version_raises(self):
        with pytest.raises(ValueError, match="Unsupported CDM version"):
            validate_patient_data({}, cdm_version="4.0")

    def test_empty_data_returns_empty(self):
        errors = validate_patient_data({}, cdm_version="5.4")
        assert errors == []

    def test_mixed_valid_and_invalid_tables(self):
        data = {
            "person": _minimal_person_df(),
            "fake_table": pl.DataFrame({"x": [1]}),
        }
        errors = validate_patient_data(data, cdm_version="5.4")
        assert len(errors) == 1
        assert "fake_table" in errors[0]

    def test_episode_table_valid_in_54_only(self):
        """episode table exists in 5.4 but not 5.3."""
        schema54 = CdmSchema(CdmVersion.V5_4)
        fields = schema54.fields_for_table("episode")
        required = [f.cdm_field_name for f in fields if f.is_required]
        data_dict: dict[str, list[object]] = {col: [1] for col in required}
        df = pl.DataFrame(data_dict)
        errors_54 = validate_patient_data({"episode": df}, cdm_version="5.4")
        assert errors_54 == []
        errors_53 = validate_patient_data({"episode": df}, cdm_version="5.3")
        assert any("Unknown CDM table" in e for e in errors_53)


# ===========================================================================
# read_patients
# ===========================================================================


class TestReadPatients:
    """Tests for read_patients()."""

    def test_read_xlsx(self, tmp_path: Path):
        tables = {
            "person": _minimal_person_df(),
            "observation_period": _minimal_obs_period_df(),
        }
        xlsx_path = tmp_path / "patients.xlsx"
        _write_test_xlsx(xlsx_path, tables)

        result = read_patients(xlsx_path, cdm_version="5.4")
        assert "person" in result
        assert "observation_period" in result
        assert len(result["person"]) == 2

    def test_read_csv_dir(self, tmp_path: Path):
        csv_dir = tmp_path / "csv_data"
        csv_dir.mkdir()
        _minimal_person_df().write_csv(csv_dir / "person.csv")
        _minimal_obs_period_df().write_csv(csv_dir / "observation_period.csv")

        result = read_patients(csv_dir, cdm_version="5.4")
        assert "person" in result
        assert "observation_period" in result

    def test_read_xlsx_with_json_output(self, tmp_path: Path):
        tables = {
            "person": _minimal_person_df(),
            "observation_period": _minimal_obs_period_df(),
        }
        xlsx_path = tmp_path / "patients.xlsx"
        _write_test_xlsx(xlsx_path, tables)
        json_out = tmp_path / "output" / "patients.json"

        read_patients(xlsx_path, cdm_version="5.4", output_path=json_out)
        assert json_out.exists()

        data = json.loads(json_out.read_text(encoding="utf-8"))
        assert "_meta" in data
        assert data["_meta"]["cdm_version"] == "5.4"
        assert "person" in data
        assert len(data["person"]) == 2

    def test_read_nonexistent_path(self):
        with pytest.raises(FileNotFoundError, match="does not exist"):
            read_patients("/nonexistent/path.xlsx")

    def test_read_unsupported_path(self, tmp_path: Path):
        txt_file = tmp_path / "data.txt"
        txt_file.write_text("not a spreadsheet")
        with pytest.raises(ValueError, match="Unsupported path"):
            read_patients(txt_file)

    def test_read_empty_csv_dir(self, tmp_path: Path):
        empty_dir = tmp_path / "empty"
        empty_dir.mkdir()
        with pytest.raises(FileNotFoundError, match="No .csv files"):
            read_patients(empty_dir)

    def test_read_invalid_data_raises(self, tmp_path: Path):
        """An xlsx with invalid table names should fail validation."""
        tables = {"bogus_table": pl.DataFrame({"x": [1]})}
        xlsx_path = tmp_path / "bad.xlsx"
        _write_test_xlsx(xlsx_path, tables)
        with pytest.raises(ValueError, match="Validation failed"):
            read_patients(xlsx_path)

    def test_read_custom_test_name(self, tmp_path: Path):
        tables = {"person": _minimal_person_df()}
        xlsx_path = tmp_path / "patients.xlsx"
        _write_test_xlsx(xlsx_path, tables)
        json_out = tmp_path / "out.json"

        read_patients(xlsx_path, test_name="my_test", output_path=json_out)
        data = json.loads(json_out.read_text(encoding="utf-8"))
        assert data["_meta"]["test_name"] == "my_test"


# ===========================================================================
# patients_cdm
# ===========================================================================


class TestPatientsCdm:
    """Tests for patients_cdm()."""

    def test_load_json(self, tmp_path: Path):
        tables = {
            "person": _minimal_person_df(),
            "observation_period": _minimal_obs_period_df(),
        }
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, tables)

        cdm = patients_cdm(json_path)
        assert isinstance(cdm, CdmReference)
        assert "person" in cdm
        assert "observation_period" in cdm
        assert cdm["person"].count() == 2

    def test_cdm_name_from_meta(self, tmp_path: Path):
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"person": _minimal_person_df()})

        cdm = patients_cdm(json_path)
        assert cdm.cdm_name == "unit_test"

    def test_cdm_name_override(self, tmp_path: Path):
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"person": _minimal_person_df()})

        cdm = patients_cdm(json_path, cdm_name="my_cdm")
        assert cdm.cdm_name == "my_cdm"

    def test_cdm_version(self, tmp_path: Path):
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"person": _minimal_person_df()})

        cdm = patients_cdm(json_path)
        assert cdm.cdm_version is CdmVersion.V5_4

    def test_tables_are_cdm_table(self, tmp_path: Path):
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"person": _minimal_person_df()})

        cdm = patients_cdm(json_path)
        assert isinstance(cdm["person"], CdmTable)

    def test_cohort_table_detected(self, tmp_path: Path):
        """A table with cohort columns should be wrapped as CohortTable."""
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"cohort": _cohort_df()})

        cdm = patients_cdm(json_path)
        assert isinstance(cdm["cohort"], CohortTable)

    def test_nonexistent_json(self):
        with pytest.raises(FileNotFoundError, match="JSON file not found"):
            patients_cdm("/nonexistent/path.json")

    def test_version_override(self, tmp_path: Path):
        # JSON meta says 5.4, but we can override at call site
        json_path = tmp_path / "test.json"
        # Write without meta
        data = {"person": [{"person_id": 1, "gender_concept_id": 8507,
                            "year_of_birth": 1990, "race_concept_id": 8515,
                            "ethnicity_concept_id": 38003563}]}
        json_path.write_text(json.dumps(data), encoding="utf-8")

        cdm = patients_cdm(json_path, cdm_version="5.3")
        assert cdm.cdm_version is CdmVersion.V5_3


# ===========================================================================
# mock_test_cdm
# ===========================================================================


class TestMockTestCdm:
    """Tests for mock_test_cdm()."""

    def test_returns_cdm_reference(self):
        cdm = mock_test_cdm()
        assert isinstance(cdm, CdmReference)

    def test_default_tables(self):
        cdm = mock_test_cdm()
        assert "person" in cdm
        assert "observation_period" in cdm
        assert "condition_occurrence" in cdm
        assert "drug_exposure" in cdm

    def test_n_persons(self):
        cdm = mock_test_cdm(n_persons=10)
        assert cdm["person"].count() == 10
        assert cdm["observation_period"].count() == 10

    def test_seed_reproducibility(self):
        cdm1 = mock_test_cdm(seed=123)
        cdm2 = mock_test_cdm(seed=123)
        df1 = cdm1["person"].collect()
        df2 = cdm2["person"].collect()
        assert df1.equals(df2)

    def test_different_seeds_differ(self):
        cdm1 = mock_test_cdm(seed=1)
        cdm2 = mock_test_cdm(seed=2)
        df1 = cdm1["person"].collect()
        df2 = cdm2["person"].collect()
        assert not df1.equals(df2)

    def test_include_conditions_false(self):
        cdm = mock_test_cdm(include_conditions=False)
        assert "condition_occurrence" not in cdm

    def test_include_drugs_false(self):
        cdm = mock_test_cdm(include_drugs=False)
        assert "drug_exposure" not in cdm

    def test_include_measurements(self):
        cdm = mock_test_cdm(include_measurements=True)
        assert "measurement" in cdm

    def test_measurements_not_included_by_default(self):
        cdm = mock_test_cdm()
        assert "measurement" not in cdm

    def test_version_53(self):
        cdm = mock_test_cdm(cdm_version="5.3")
        assert cdm.cdm_version is CdmVersion.V5_3

    def test_version_54(self):
        cdm = mock_test_cdm(cdm_version="5.4")
        assert cdm.cdm_version is CdmVersion.V5_4

    def test_person_columns(self):
        cdm = mock_test_cdm()
        cols = set(cdm["person"].columns)
        assert "person_id" in cols
        assert "gender_concept_id" in cols
        assert "year_of_birth" in cols

    def test_observation_period_columns(self):
        cdm = mock_test_cdm()
        cols = set(cdm["observation_period"].columns)
        assert "observation_period_id" in cols
        assert "person_id" in cols
        assert "observation_period_start_date" in cols
        assert "observation_period_end_date" in cols

    def test_cdm_name(self):
        cdm = mock_test_cdm()
        assert cdm.cdm_name == "mock_test"

    def test_single_person(self):
        cdm = mock_test_cdm(n_persons=1)
        assert cdm["person"].count() == 1


# ===========================================================================
# generate_test_tables
# ===========================================================================


class TestGenerateTestTables:
    """Tests for generate_test_tables()."""

    def test_creates_xlsx(self, tmp_path: Path):
        path = generate_test_tables(
            ["person", "observation_period"],
            output_path=tmp_path,
        )
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_correct_sheets(self, tmp_path: Path):
        import openpyxl

        path = generate_test_tables(
            ["person", "observation_period"],
            output_path=tmp_path,
        )
        wb = openpyxl.load_workbook(path, read_only=True)
        sheet_names = set(wb.sheetnames)
        wb.close()
        assert sheet_names == {"person", "observation_period"}

    def test_correct_columns(self, tmp_path: Path):
        import openpyxl

        path = generate_test_tables(["person"], output_path=tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["person"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()

        schema = CdmSchema(CdmVersion.V5_4)
        expected = [f.cdm_field_name for f in schema.fields_for_table("person")]
        assert headers == expected

    def test_custom_filename(self, tmp_path: Path):
        path = generate_test_tables(
            ["person"],
            output_path=tmp_path,
            filename="my_template.xlsx",
        )
        assert path.name == "my_template.xlsx"

    def test_default_filename(self, tmp_path: Path):
        path = generate_test_tables(["person"], output_path=tmp_path)
        assert "v54" in path.name

    def test_invalid_table_raises(self, tmp_path: Path):
        with pytest.raises(ValueError, match="Unknown CDM table"):
            generate_test_tables(["bogus_table"], output_path=tmp_path)

    def test_vocab_table_rejected(self, tmp_path: Path):
        with pytest.raises(ValueError, match="Vocabulary table"):
            generate_test_tables(["concept"], output_path=tmp_path)

    def test_version_53(self, tmp_path: Path):
        path = generate_test_tables(
            ["person"], output_path=tmp_path, cdm_version="5.3"
        )
        assert path.exists()

    def test_multiple_tables(self, tmp_path: Path):
        import openpyxl

        tables = ["person", "observation_period", "condition_occurrence", "drug_exposure"]
        path = generate_test_tables(tables, output_path=tmp_path)
        wb = openpyxl.load_workbook(path, read_only=True)
        assert set(wb.sheetnames) == set(tables)
        wb.close()


# ===========================================================================
# graph_cohort
# ===========================================================================


class TestGraphCohort:
    """Tests for graph_cohort()."""

    def test_returns_plotly_figure(self):
        import plotly.graph_objects as go

        fig = graph_cohort(
            subject_id=1,
            cohorts={"target": _cohort_df()},
        )
        assert isinstance(fig, go.Figure)

    def test_single_cohort(self):
        fig = graph_cohort(subject_id=1, cohorts={"target": _cohort_df()})
        # Subject 1 has one entry in cohort def 1
        assert len(fig.data) > 0

    def test_multiple_cohorts(self):
        fig = graph_cohort(
            subject_id=1,
            cohorts={
                "target": _cohort_df(),
                "comparator": _cohort_df(),
            },
        )
        assert len(fig.data) > 0

    def test_no_records_raises(self):
        df = _cohort_df().filter(pl.col("subject_id") == 999)
        with pytest.raises(ValueError, match="No cohort records"):
            graph_cohort(subject_id=1, cohorts={"target": df})

    def test_missing_columns_raises(self):
        bad_df = pl.DataFrame({"subject_id": [1], "x": [1]})
        with pytest.raises(ValueError, match="missing required columns"):
            graph_cohort(subject_id=1, cohorts={"target": bad_df})

    def test_style_override(self):
        fig = graph_cohort(
            subject_id=1,
            cohorts={"target": _cohort_df()},
            style={"title": "Custom Title"},
        )
        assert fig.layout.title.text == "Custom Title"


# ===========================================================================
# JSON round-trip
# ===========================================================================


class TestJsonRoundTrip:
    """Tests for end-to-end JSON round-trip."""

    def test_xlsx_to_json_to_cdm(self, tmp_path: Path):
        tables = {
            "person": _minimal_person_df(),
            "observation_period": _minimal_obs_period_df(),
        }
        xlsx_path = tmp_path / "patients.xlsx"
        json_path = tmp_path / "patients.json"
        _write_test_xlsx(xlsx_path, tables)

        # Read from xlsx, write to JSON
        read_patients(xlsx_path, output_path=json_path)

        # Load from JSON into CDM
        cdm = patients_cdm(json_path)
        assert isinstance(cdm, CdmReference)
        assert cdm["person"].count() == 2
        assert cdm["observation_period"].count() == 2

    def test_json_round_trip_preserves_data(self, tmp_path: Path):
        """Data values should survive the JSON round-trip."""
        original = _minimal_person_df()
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, {"person": original})

        cdm = patients_cdm(json_path)
        loaded = cdm["person"].collect()
        assert loaded["person_id"].to_list() == original["person_id"].to_list()
        assert loaded["gender_concept_id"].to_list() == original["gender_concept_id"].to_list()

    def test_create_then_read_json(self, tmp_path: Path):
        """Write a JSON, read it back as raw Python, verify structure."""
        tables = {"person": _minimal_person_df()}
        json_path = tmp_path / "test.json"
        _write_test_json(json_path, tables)

        raw = json.loads(json_path.read_text(encoding="utf-8"))
        assert "_meta" in raw
        assert "person" in raw
        assert isinstance(raw["person"], list)
        assert len(raw["person"]) == 2


# ===========================================================================
# Integration / edge cases
# ===========================================================================


class TestEdgeCases:
    """Edge cases and integration tests."""

    def test_mock_cdm_then_validate(self):
        """mock_test_cdm should produce data that passes validate_patient_data."""
        cdm = mock_test_cdm(n_persons=3)
        data = {}
        for name in cdm.table_names:
            tbl = cdm[name]
            data[name] = tbl.collect()
        errors = validate_patient_data(data, cdm_version="5.4")
        assert errors == []

    def test_generate_then_read(self, tmp_path: Path):
        """generate_test_tables -> fill in data -> read_patients."""
        # Generate template
        path = generate_test_tables(["person"], output_path=tmp_path)

        # Fill in data using openpyxl
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb["person"]
        # Find column indices for required fields
        headers = [cell.value for cell in ws[1]]
        pid_idx = headers.index("person_id") + 1
        gender_idx = headers.index("gender_concept_id") + 1
        yob_idx = headers.index("year_of_birth") + 1
        race_idx = headers.index("race_concept_id") + 1
        eth_idx = headers.index("ethnicity_concept_id") + 1

        ws.cell(row=2, column=pid_idx, value=1)
        ws.cell(row=2, column=gender_idx, value=8507)
        ws.cell(row=2, column=yob_idx, value=1990)
        ws.cell(row=2, column=race_idx, value=8515)
        ws.cell(row=2, column=eth_idx, value=38003563)
        wb.save(str(path))
        wb.close()

        # Now read_patients should work
        result = read_patients(path, cdm_version="5.4")
        assert "person" in result
        assert len(result["person"]) == 1
