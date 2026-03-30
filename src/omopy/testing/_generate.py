"""Generate blank Excel templates with CDM table schemas."""

from __future__ import annotations

from pathlib import Path

from omopy.generics import CdmSchema, CdmVersion, TableGroup

__all__ = ["generate_test_tables"]


# ---------------------------------------------------------------------------
# Version helpers
# ---------------------------------------------------------------------------

_CDM_VERSION_MAP: dict[str, CdmVersion] = {
    "5.3": CdmVersion.V5_3,
    "5.4": CdmVersion.V5_4,
}


def _resolve_version(cdm_version: str) -> CdmVersion:
    try:
        return _CDM_VERSION_MAP[cdm_version]
    except KeyError:
        msg = f"Unsupported CDM version {cdm_version!r}. Supported: {list(_CDM_VERSION_MAP)}"
        raise ValueError(msg) from None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_test_tables(
    table_names: list[str],
    *,
    cdm_version: str = "5.4",
    output_path: str | Path = ".",
    filename: str | None = None,
) -> Path:
    """Generate an empty Excel file with sheets for specified CDM tables.

    Each sheet contains the correct column headers from the CDM
    specification. Vocabulary tables (``concept``, ``concept_ancestor``,
    etc.) are excluded automatically.

    Args:
        table_names: List of CDM table names to include as sheets.
        cdm_version: CDM version string (``"5.3"`` or ``"5.4"``).
        output_path: Directory where the file will be created.
        filename: Output filename. Defaults to
            ``"test_patients_v{version}.xlsx"``.

    Returns:
        Path to the created Excel file.

    Raises:
        ValueError: If any table name is not a valid CDM table, or if
            a vocabulary table is requested.
    """
    import openpyxl

    version = _resolve_version(cdm_version)
    schema = CdmSchema(version)
    valid_tables = set(schema.table_names())
    vocab_tables = set(schema.table_names_in_group(TableGroup.VOCAB))

    # Validate requested tables
    errors: list[str] = []
    for name in table_names:
        if name not in valid_tables:
            errors.append(f"Unknown CDM table: '{name}'")
        elif name in vocab_tables:
            errors.append(
                f"Vocabulary table '{name}' excluded from test templates "
                "(use a real database for vocabulary data)"
            )
    if errors:
        msg = "Invalid table names:\n" + "\n".join(f"  - {e}" for e in errors)
        raise ValueError(msg)

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    for table_name in table_names:
        fields = schema.fields_for_table(table_name)
        ws = wb.create_sheet(title=table_name)
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field.cdm_field_name)
            # Bold required fields
            if field.is_required:
                cell.font = openpyxl.styles.Font(bold=True)

    # Write file
    out_dir = Path(output_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = filename or f"test_patients_v{cdm_version.replace('.', '')}.xlsx"
    out_file = out_dir / fname
    wb.save(str(out_file))
    wb.close()

    return out_file
