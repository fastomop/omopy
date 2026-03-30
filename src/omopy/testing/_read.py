"""Patient data reading and validation against the OMOP CDM spec."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import polars as pl

from omopy.generics import CdmSchema, CdmVersion, TableGroup

__all__ = ["read_patients", "validate_patient_data"]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_CDM_VERSION_MAP: dict[str, CdmVersion] = {
    "5.3": CdmVersion.V5_3,
    "5.4": CdmVersion.V5_4,
}


def _resolve_version(cdm_version: str) -> CdmVersion:
    """Convert a version string to a ``CdmVersion`` enum."""
    try:
        return _CDM_VERSION_MAP[cdm_version]
    except KeyError:
        msg = f"Unsupported CDM version {cdm_version!r}. Supported: {list(_CDM_VERSION_MAP)}"
        raise ValueError(msg) from None


def _df_to_serializable(df: pl.DataFrame) -> list[dict[str, Any]]:
    """Convert a Polars DataFrame to a JSON-serializable list of dicts.

    Handles date/datetime columns by converting them to ISO 8601 strings.
    """
    rows: list[dict[str, Any]] = []
    for row in df.iter_rows(named=True):
        converted: dict[str, Any] = {}
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                converted[k] = v.isoformat()
            else:
                converted[k] = v
        rows.append(converted)
    return rows


def _read_xlsx(path: Path) -> dict[str, pl.DataFrame]:
    """Read an Excel workbook: each sheet becomes a table."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, pl.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h is not None else f"_col{i}" for i, h in enumerate(rows[0])]
        if len(rows) > 1:
            data_rows = rows[1:]
        else:
            data_rows = []
        # Build column-oriented dict
        col_data: dict[str, list[Any]] = {h: [] for h in headers}
        for row in data_rows:
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                col_data[h].append(val)
        result[sheet_name.lower()] = pl.DataFrame(col_data)
    wb.close()
    return result


def _read_csv_dir(path: Path) -> dict[str, pl.DataFrame]:
    """Read a directory of CSV files: each file becomes a table."""
    result: dict[str, pl.DataFrame] = {}
    for csv_file in sorted(path.glob("*.csv")):
        table_name = csv_file.stem.lower()
        result[table_name] = pl.read_csv(csv_file, infer_schema_length=1000)
    if not result:
        msg = f"No .csv files found in directory: {path}"
        raise FileNotFoundError(msg)
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def validate_patient_data(
    data: dict[str, pl.DataFrame],
    *,
    cdm_version: str = "5.4",
) -> list[str]:
    """Validate patient data against the OMOP CDM specification.

    Checks that each table name is a valid CDM table, that column names
    match the CDM field specs, and that required fields are present.

    Args:
        data: Mapping of table name to Polars DataFrame.
        cdm_version: CDM version string (``"5.3"`` or ``"5.4"``).

    Returns:
        A list of error messages. An empty list means the data is valid.
    """
    version = _resolve_version(cdm_version)
    schema = CdmSchema(version)
    valid_tables = set(schema.table_names())
    errors: list[str] = []

    for table_name, df in data.items():
        if table_name not in valid_tables:
            errors.append(f"Unknown CDM table: '{table_name}'")
            continue

        # Check columns against spec
        col_errors = schema.validate_columns(table_name, df.columns, check_required=True)
        errors.extend(col_errors)

    return errors


def read_patients(
    path: str | Path,
    *,
    cdm_version: str = "5.4",
    test_name: str = "test",
    output_path: str | Path | None = None,
) -> dict[str, pl.DataFrame]:
    """Read patient data from an Excel file or CSV directory.

    Auto-detects the format based on the path:

    * If ``path`` ends with ``.xlsx``, reads each sheet as a CDM table
      (sheet name = table name).
    * If ``path`` is a directory, reads each ``.csv`` file as a CDM table
      (filename stem = table name).

    The data is validated against the CDM specification. If validation
    fails, a ``ValueError`` is raised with all error messages.

    If ``output_path`` is provided, writes the data as a JSON file
    (format: ``{"table_name": [{col: val, ...}, ...], ...}``).

    Args:
        path: Path to an ``.xlsx`` file or a directory of ``.csv`` files.
        cdm_version: CDM version string (``"5.3"`` or ``"5.4"``).
        test_name: Name for this test patient set (used in JSON metadata).
        output_path: Optional path to write JSON output.

    Returns:
        A dict mapping table names to Polars DataFrames.

    Raises:
        ValueError: If the data fails CDM validation.
        FileNotFoundError: If the path does not exist or contains no data.
    """
    p = Path(path)
    if not p.exists():
        msg = f"Path does not exist: {p}"
        raise FileNotFoundError(msg)

    if p.suffix.lower() == ".xlsx":
        data = _read_xlsx(p)
    elif p.is_dir():
        data = _read_csv_dir(p)
    else:
        msg = f"Unsupported path: {p}. Provide an .xlsx file or a directory of .csv files."
        raise ValueError(msg)

    if not data:
        msg = f"No tables found at path: {p}"
        raise ValueError(msg)

    # Validate
    errors = validate_patient_data(data, cdm_version=cdm_version)
    if errors:
        msg = f"Validation failed with {len(errors)} error(s):\n" + "\n".join(
            f"  - {e}" for e in errors
        )
        raise ValueError(msg)

    # Optionally write JSON
    if output_path is not None:
        out_p = Path(output_path)
        json_data: dict[str, Any] = {
            "_meta": {"test_name": test_name, "cdm_version": cdm_version},
        }
        for tbl_name, df in data.items():
            json_data[tbl_name] = _df_to_serializable(df)
        out_p.parent.mkdir(parents=True, exist_ok=True)
        out_p.write_text(json.dumps(json_data, indent=2, default=str), encoding="utf-8")

    return data
